"""
9기 수기집 <항해일지> 평가 스크립트 (OpenAI 버전)
- 각 수기를 OpenAI API로 평가하여 4가지 지표 점수(0~5), 최종 점수, 선별 여부, AI 사용도, 한 줄 평을 산출
- 결과는 '_evaluated.xlsx'로 저장 (resume 지원)
- OPENAI_API_KEY 환경변수 필요
"""
from __future__ import annotations

import argparse
import json
import os
import re
import sys
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path

import openpyxl
from openai import OpenAI, APIError

MODEL = os.getenv("OPENAI_MODEL", "gpt-4o")  # gpt-4o (강력), gpt-4o-mini (저렴), gpt-5 (사용 가능시)
INPUT_PATH = Path(__file__).parent / "9기 수기 정리_260521_1738.xlsx"
OUTPUT_PATH = Path(__file__).parent / "9기 수기 정리_260521_1738_evaluated.xlsx"
SHEET_NAME = "수기정리"

COL_NAME = 1
COL_APPLY_TYPE = 2
COL_UNIV = 3
COL_ITEM = 4
COL_CHARS = 5
COL_BODY = 8
COL_FINAL_SCORE = 14
COL_SELECTED = 15
COL_AI_USAGE = 16
COL_ONELINE = 17

SAVE_EVERY = 20
MAX_WORKERS = 8
MAX_RETRIES = 4

SYSTEM_PROMPT = """당신은 대한민국 최상위권 입시 학원 '시대인재'의 공식 수기집 <항해일지> 발간을 담당하는 수석 AI 평가 에이전트입니다.
입력된 학생의 수기를 면밀히 분석하여 4가지 핵심 평가 지표를 기준으로 각각 0~5점의 정수 점수를 부여하고, 추가로 AI 사용 의심도와 한 줄 평을 작성해 JSON 형태로 반환합니다.

<evaluation_criteria>
<criterion id="1" weight="최우선">
<name>정성과 구체성 (Sincerity & Specificity)</name>
<focus>답변의 분량이 충분하며, 뜬구름 잡는 소리가 아닌 구체적인 예시와 데이터를 포함하는가?</focus>
<positive_indicators>특정 과목 단원명, 특정 시기(예: 6평 직후), 모의고사 점수 변화, 본인만의 문제 풀이 플로우 등 상세 서술</positive_indicators>
<negative_indicators>단답형, "기출이 중요하다" 수준의 추상적이고 보편적인 문장, 성의 없는 짧은 분량</negative_indicators>
</criterion>

<criterion id="2">
<name>실용성과 타겟 적합성 (Utility & Actionable Advice)</name>
<focus>다음 기수 시대인재N 재원생들이 즉시 적용 가능한 '행동 영역(Action Item)'이 존재하는가?</focus>
<positive_indicators>시대인재 전용 콘텐츠(서바이벌, 엣지 등) 활용법, 실전 시간 관리법 등</positive_indicators>
<negative_indicators>개인의 특출난 재능에만 의존한 방법론, 타인이 모방하기 힘든 예외적 상황</negative_indicators>
</criterion>

<criterion id="3">
<name>내용의 독창성과 희소성 (Uniqueness)</name>
<focus>흔한 합격 수기의 클리셰를 벗어난 학생만의 치열한 고민이나 특별한 서사가 있는가?</focus>
<positive_indicators>독특한 멘탈 관리법, 선택과목 변경 극복기, 반수생 적응기 등 참신한 케이스</positive_indicators>
<negative_indicators>인터넷 커뮤니티 등에서 쉽게 볼 수 있는 뻔한 내용 답습</negative_indicators>
</criterion>

<criterion id="4">
<name>가독성과 스토리텔링 (Readability)</name>
<focus>글의 흐름이 매끄럽고 독자(수험생)가 몰입해서 읽을 수 있는가?</focus>
<positive_indicators>위기-극복-성취의 기승전결 구조, 수능 당일의 현장감/감정선 묘사, 오류 없는 문법</positive_indicators>
<negative_indicators>잦은 비문, 뚝뚝 끊기는 문맥, 지나치게 감정적인 어휘 남발</negative_indicators>
</criterion>
</evaluation_criteria>

<scoring_scale>
5: 최우수 (발췌 1순위) — 구체적 예시·탁월한 실용성·완벽한 문장력 모두 갖춤. 즉시 원고로 사용 가능.
4: 우수 (적극 활용 가능) — 정성과 실질적인 팁이 명확. 약간의 윤문만 거치면 훌륭한 원고.
3: 보통 (내용 무난/중복 가능성) — 내용은 무난하나 다소 진부하거나 추상적. 부분적 발췌만 가능.
2: 미흡 (활용도 낮음) — 구체성 부족, 빈약하거나 감정적 서술에 치우쳐 실질적 정보 부족.
1: 매우 미흡 (분량 부족/무성의) — 단답형이거나 분량 극히 적어 정성을 찾기 힘듦.
0: 활용 불가 — 질문 의도와 무관한 답변, 비속어 포함 등 가치 전무.
</scoring_scale>

<output_format>
오직 아래 JSON 한 객체만 반환합니다. 코드펜스/설명/추가 텍스트 금지.
{
  "정성_구체성": <int 0-5>,
  "실용성_타겟적합성": <int 0-5>,
  "독창성": <int 0-5>,
  "가독성": <int 0-5>,
  "최종_점수": <int 0-5, 네 지표 산술평균을 반올림한 정수>,
  "선별_여부": <bool, 최종_점수 >= 4 이면 true>,
  "AI_사용도": <int 0-5, 0=확실히 사람이 직접 작성, 5=거의 확실히 AI 생성. 시대인재 강사명/콘텐츠명/구체적 시기 등이 등장하면 사람일 확률이 높음>,
  "한줄평": "<1~2문장의 평가 사유. 강점과 약점을 간결히>"
}
</output_format>"""


def build_user_message(row: dict) -> str:
    return (
        "<journal>\n"
        f"이름: {row['이름']}\n"
        f"지원 형태: {row.get('지원_형태', '')}\n"
        f"합격 대학: {row.get('합격_대학', '')}\n"
        f"수기 작성 항목: {row.get('항목', '')}\n"
        f"글자수(공백포함): {row.get('글자수', '')}\n"
        "본문:\n"
        f"{row['본문']}\n"
        "</journal>\n\n"
        "위 수기를 평가하여 지정된 JSON만 반환하세요."
    )


def parse_json_response(text: str) -> dict:
    text = text.strip()
    text = re.sub(r"^```(?:json)?\s*", "", text)
    text = re.sub(r"\s*```$", "", text)
    m = re.search(r"\{.*\}", text, re.DOTALL)
    if not m:
        raise ValueError(f"No JSON object in response: {text[:200]}")
    return json.loads(m.group(0))


def coerce_int(v, lo=0, hi=5) -> int:
    if isinstance(v, bool):
        return int(v)
    if isinstance(v, (int, float)):
        return max(lo, min(hi, int(round(v))))
    if isinstance(v, str):
        try:
            return max(lo, min(hi, int(round(float(v)))))
        except ValueError:
            pass
    raise ValueError(f"cannot coerce to int: {v!r}")


def coerce_bool(v) -> bool:
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        return bool(v)
    if isinstance(v, str):
        return v.strip().lower() in ("true", "1", "yes", "y", "t", "o")
    return False


def evaluate_one(client: OpenAI, row: dict) -> dict:
    user_msg = build_user_message(row)
    last_err = None
    for attempt in range(MAX_RETRIES):
        try:
            resp = client.chat.completions.create(
                model=MODEL,
                messages=[
                    {"role": "system", "content": SYSTEM_PROMPT},
                    {"role": "user", "content": user_msg},
                ],
                response_format={"type": "json_object"},
                temperature=0.3,
                max_tokens=512,
            )
            text = resp.choices[0].message.content
            data = parse_json_response(text)
            out = {
                "정성_구체성": coerce_int(data.get("정성_구체성", 0)),
                "실용성_타겟적합성": coerce_int(data.get("실용성_타겟적합성", 0)),
                "독창성": coerce_int(data.get("독창성", 0)),
                "가독성": coerce_int(data.get("가독성", 0)),
                "AI_사용도": coerce_int(data.get("AI_사용도", 0)),
                "한줄평": str(data.get("한줄평", "")).strip(),
            }
            avg = (out["정성_구체성"] + out["실용성_타겟적합성"] + out["독창성"] + out["가독성"]) / 4
            out["최종_점수"] = max(0, min(5, int(round(avg))))
            if "최종_점수" in data:
                try:
                    out["최종_점수"] = coerce_int(data["최종_점수"])
                except Exception:
                    pass
            out["선별_여부"] = coerce_bool(data.get("선별_여부", out["최종_점수"] >= 4))
            out["_usage"] = {
                "input_tokens": resp.usage.prompt_tokens,
                "output_tokens": resp.usage.completion_tokens,
            }
            return out
        except (APIError, ValueError, json.JSONDecodeError) as e:
            last_err = e
            wait = 2 ** attempt
            print(f"  [retry {attempt+1}/{MAX_RETRIES}] row={row['row']} err={type(e).__name__}: {str(e)[:200]} (sleep {wait}s)", file=sys.stderr)
            time.sleep(wait)
    raise RuntimeError(f"Failed after {MAX_RETRIES} retries: {last_err}")


def load_jobs(ws, start_row=2, end_row=None, force=False):
    end_row = end_row or ws.max_row
    jobs = []
    for r in range(start_row, end_row + 1):
        body = ws.cell(r, COL_BODY).value
        if not body or not str(body).strip():
            continue
        if not force and ws.cell(r, COL_FINAL_SCORE).value is not None:
            continue
        jobs.append({
            "row": r,
            "이름": ws.cell(r, COL_NAME).value or "",
            "지원_형태": ws.cell(r, COL_APPLY_TYPE).value or "",
            "합격_대학": ws.cell(r, COL_UNIV).value or "",
            "항목": ws.cell(r, COL_ITEM).value or "",
            "글자수": ws.cell(r, COL_CHARS).value or "",
            "본문": str(body),
        })
    return jobs


def write_result(ws, row_idx: int, result: dict):
    ws.cell(row_idx, COL_FINAL_SCORE).value = result["최종_점수"]
    ws.cell(row_idx, COL_SELECTED).value = "O" if result["선별_여부"] else "X"
    ws.cell(row_idx, COL_AI_USAGE).value = result["AI_사용도"]
    ws.cell(row_idx, COL_ONELINE).value = result["한줄평"]


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--start", type=int, default=2)
    parser.add_argument("--end", type=int, default=None)
    parser.add_argument("--limit", type=int, default=None)
    parser.add_argument("--force", action="store_true")
    parser.add_argument("--workers", type=int, default=MAX_WORKERS)
    parser.add_argument("--input", type=str, default=str(INPUT_PATH))
    parser.add_argument("--output", type=str, default=str(OUTPUT_PATH))
    args = parser.parse_args()

    if not os.getenv("OPENAI_API_KEY"):
        print("ERROR: OPENAI_API_KEY 환경변수가 필요합니다.", file=sys.stderr)
        sys.exit(2)

    src = Path(args.output) if Path(args.output).exists() else Path(args.input)
    print(f"Loading: {src}")
    wb = openpyxl.load_workbook(src)
    ws = wb[SHEET_NAME]

    jobs = load_jobs(ws, args.start, args.end, force=args.force)
    if args.limit:
        jobs = jobs[: args.limit]
    print(f"Jobs to evaluate: {len(jobs)} (workers={args.workers}, model={MODEL})")
    if not jobs:
        print("Nothing to do.")
        return

    client = OpenAI()
    completed = 0
    failed = 0
    total_in = total_out = 0
    t0 = time.time()
    save_counter = 0

    with ThreadPoolExecutor(max_workers=args.workers) as ex:
        futures = {ex.submit(evaluate_one, client, job): job for job in jobs}
        for fut in as_completed(futures):
            job = futures[fut]
            try:
                result = fut.result()
                write_result(ws, job["row"], result)
                usage = result.get("_usage", {})
                total_in += usage.get("input_tokens", 0)
                total_out += usage.get("output_tokens", 0)
                completed += 1
                print(
                    f"[{completed}/{len(jobs)}] row={job['row']} {str(job['이름'])[:8]:<8} "
                    f"score={result['최종_점수']} sel={'O' if result['선별_여부'] else 'X'} "
                    f"ai={result['AI_사용도']} | {result['한줄평'][:60]}"
                )
                save_counter += 1
                if save_counter >= SAVE_EVERY:
                    wb.save(args.output)
                    save_counter = 0
                    print(f"  -- saved checkpoint ({completed}/{len(jobs)})")
            except Exception as e:
                failed += 1
                print(f"FAILED row={job['row']} ({job['이름']}): {e}", file=sys.stderr)

    wb.save(args.output)
    dt = time.time() - t0
    print(f"\nDone. completed={completed} failed={failed} elapsed={dt:.1f}s")
    print(f"Tokens: in={total_in} out={total_out} model={MODEL}")
    print(f"Saved to: {args.output}")


if __name__ == "__main__":
    main()
